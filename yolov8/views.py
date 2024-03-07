import ast
import json
import multiprocessing
import os
import string
from collections import Counter
from datetime import datetime
from pathlib import Path
import random

import cv2
import thread6
from django.http import JsonResponse, HttpResponse, FileResponse, Http404
from django.shortcuts import render, redirect, get_object_or_404
from ultralytics import YOLO
from openpyxl import Workbook

from users.forms import RegistrationForm
from .forms import VideoForm
from .models import Video, Report

base_dir = Path(__file__).resolve().parent


def video_list(request):
    if request.method == 'POST':
        form = VideoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('yolov8:video_list')
    else:
        form = VideoForm()
    videos = Video.objects.all()
    register_form = RegistrationForm()
    return render(request, 'video_list.html',
                  {'videos': videos, 'form': form, 'user': request.user, 'register_form': register_form})


def upload_video(request):
    if request.method == 'POST':
        form = VideoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('yolov8:video_list')
    else:
        form = VideoForm()
    return render(request, 'video_list.html', {'form': form})


def analyze_video(request):
    params = request.GET
    # model_pt = params.get('model')
    model_pt = "best.pt"
    video_name = params.get('video')
    video_id = params.get('id')
    video_name = video_name.split('/')[-1]
    video_analysis_mul_process(model_pt, video_name, video_id, request)
    videos = Video.objects.all()
    return render(request, 'video_list.html', {'videos': videos})


@thread6.threaded(False)
def video_analysis_mul_process(model_pt, video_name, video_id, request):
    lock = multiprocessing.Semaphore(2)
    p = multiprocessing.Process(target=video_analysis,
                                args=(lock, model_pt, video_name, video_id, request))
    p.run()


def video_analysis(lock, model_pt, video_name, video_id, request):
    lock.acquire()
    model_name = str(base_dir) + "\\config\\" + model_pt
    video = str(base_dir) + "\\config\\videos\\" + video_name
    model = YOLO(model_name)
    cap = cv2.VideoCapture(video)
    results = []
    total = 0
    while True:
        total += 1
        ret, im2 = cap.read()
        if not ret:
            # cv2.waitKey(0)
            break
        im2 = cv2.resize(im2, (0, 0), fx=0.5, fy=0.5)
        results.append(model.predict(source=im2, save=True, save_txt=True, show=True,
                                     save_crop=True, ))  # save predictions as labelsre
    counter = analyze_result(results)
    video_obj = Video.objects.get(id=int(video_id))
    user = request.user
    report_write2file = counter
    report_write2file['用户'] = request.user.username
    report_write2file['总帧数'] = total
    report_write2file['检测时间'] = gen_date_time()
    report_write2file_lst = [report_write2file]
    file_name_origin = gen_file_name()
    file_name = str(base_dir) + "\\config\\report\\" + file_name_origin
    create_excel_from_dict(report_write2file_lst, file_name)
    report = Report(video=video_obj,
                    report=counter,
                    total=total,
                    user=user,
                    file_name=file_name_origin)
    report.save()
    lock.release()


def analyze_result(results):
    filter_type = list(results[0][0].names.keys())
    filter_type_human = {"D00": "垂直裂纹帧数",
                         "D10": "沥青路面疲劳裂缝帧数",
                         "D20": "沥青路面反射裂缝帧数",
                         "D30": "水泥混凝土路面龟裂帧数",
                         "D40": "路面结构沉降帧数",
                         "D44": "路基沉降帧数",
                         "D50": "沥青路面车辙帧数",
                         "D01": "路面表层材料松散帧数",
                         "D11": "沥青路面坑洞帧数",
                         "D43": "水泥混凝土路面坑洞帧数",
                         "D0w0": "路面碎屑帧数",
                         "D81": "路面颗粒状损坏帧数"}
    all_results = []
    for result in results:
        if result[0].boxes.cls.numel() > 0:
            all_results.extend(result[0].boxes.cls.tolist())
    all_results = [int(x) for x in all_results]
    counter = Counter(all_results)
    try:
        result_dict = {filter_type_human[results[0][0].names[key]]: counter[key] for key in filter_type}
        return result_dict
    except IndexError:
        print("提交的视频有误")
        pass


def create_excel_from_dict(data, file_name):
    # 创建一个工作簿对象
    wb = Workbook()
    # 激活默认的工作表
    ws = wb.active

    # 提取字典的键并将其写入第一行
    headers = list(data[0].keys())
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header)

    # 将字典数据写入Excel文件
    for row_num, row_data in enumerate(data, start=2):
        for col_num, cell_value in enumerate(row_data.values(), start=1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # 保存Excel文件
    wb.save(file_name)


def upload_report(video_name, report):
    pass


def delete_video(request, video_id):
    if request.method == 'POST':
        video = get_object_or_404(Video, pk=video_id)
        video_file = str(base_dir) + "\\config\\" + str(video.file.name)
        delete_file(video_file)
        video.delete()
        return JsonResponse({'message': 'Video deleted successfully'})
    else:
        return JsonResponse({'error': 'Invalid request method'})


def delete_file(file_path):
    try:
        os.remove(file_path)
        print(f"File '{file_path}' deleted successfully.")
    except OSError as e:
        print(f"Error deleting file '{file_path}': {e}")


def check_report(request, video_id):
    # 根据 video_id 查询对应的视频
    video = get_object_or_404(Video, id=int(video_id))

    # 根据视频查询报告
    report = Report.objects.filter(video=video).first()

    # 如果报告存在，则返回 True；否则返回 False
    has_report = True if report else False

    # 返回 JSON 响应
    return JsonResponse({'has_report': has_report})


def show_report(request, video_id):
    # 根据 video_id 查询对应的视频
    video = get_object_or_404(Video, id=int(video_id))

    # 根据视频查询报告
    report_obj = Report.objects.filter(video=video).first()
    total = report_obj.total
    report_result = json.loads(json.dumps(report_obj.report))
    report_result = ast.literal_eval(report_result)
    report_result["总帧数"] = total

    return JsonResponse({'report': report_result})


def download_report(request, video_id):
    # 根据 video_id 查询对应的视频
    video = get_object_or_404(Video, id=int(video_id))

    # 根据视频查询报告
    report_obj = Report.objects.filter(video=video).order_by('-created_at').first()
    report_file_name = report_obj.file_name
    file_path = str(base_dir) + "\\config\\report\\" + report_file_name
    # 检查文件是否存在
    if os.path.exists(file_path):
        # 返回报告文件作为 HTTP 响应
        return FileResponse(open(file_path, 'rb'),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # 如果文件不存在，则返回404错误
    raise Http404("报告文件未找到")


def gen_date_time():
    # 获取当前日期和时间
    current_time = datetime.now()
    # 格式化当前时间为字符串
    formatted_time = current_time.strftime("%Y-%m-%d %H:%M:%S")

    return formatted_time


def gen_file_name(length=8):
    # 生成包含大小写字母和数字的字符集
    characters = string.ascii_letters + string.digits
    # 从字符集中随机选择length个字符，并拼接成字符串
    random_string = ''.join(random.choice(characters) for _ in range(length))
    return random_string + ".xlsx"
